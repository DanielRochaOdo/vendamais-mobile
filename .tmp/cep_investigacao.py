import json, re, urllib.parse, urllib.request
from collections import defaultdict
from openpyxl import load_workbook

xlsx_path = r"c:\Users\daniel.rocha\Desktop\falta de ceps s4e.xlsx"
SUPABASE_URL = "https://plonbokgcxwsdqfyjkwl.supabase.co"
SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InBsb25ib2tnY3h3c2RxZnlqa3dsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2NjA3Njg5OSwiZXhwIjoyMDgxNjUyODk5fQ.XTi7ZOVkj8Y7AqcA1plvpPP0W3NYDfX47_GWwdhTrUE"
headers = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Accept": "application/json"
}

wb = load_workbook(xlsx_path, data_only=True)
ws = wb["Planilha1"]
header = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
if "cpf titular" not in header:
    raise SystemExit("Coluna 'cpf titular' não encontrada em Planilha1")
cpf_idx = header.index("cpf titular") + 1

raw_cpfs = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, cpf_idx).value
    if v is None:
        continue
    s = re.sub(r"\D", "", str(v))
    if not s:
        continue
    s = s.zfill(11)
    if len(s) > 11:
        s = s[-11:]
    raw_cpfs.append(s)
cpfs = sorted(set(raw_cpfs))

def http_get_json(url, query):
    qs = urllib.parse.urlencode(query, safe='(),>')
    req = urllib.request.Request(url + '?' + qs, headers=headers, method='GET')
    with urllib.request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read().decode('utf-8'))

def fetch_table(table, params):
    out = []
    offset = 0
    limit = 1000
    while True:
        p = params.copy()
        p["limit"] = str(limit)
        p["offset"] = str(offset)
        rows = http_get_json(f"{SUPABASE_URL}/rest/v1/{table}", p)
        out.extend(rows)
        if len(rows) < limit:
            break
        offset += limit
    return out

cadastros = []
for i in range(0, len(cpfs), 150):
    batch = cpfs[i:i+150]
    in_expr = '(' + ','.join(batch) + ')'
    cadastros.extend(fetch_table("cadastros", {
        "select": "id,cpf,created_at,status,tipo_cadastro,endereco,payload_erp",
        "cpf": f"in.{in_expr}",
        "order": "created_at.asc"
    }))

cad_by_cpf = defaultdict(list)
cad_ids = set()
for c in cadastros:
    ccpf = re.sub(r"\D", "", str(c.get("cpf") or "")).zfill(11)[-11:]
    cad_by_cpf[ccpf].append(c)
    cad_ids.add(str(c.get("id")))

logs = fetch_table("api_logs", {
    "select": "id,created_at,endpoint,request_body,status_code",
    "endpoint": "eq.erp-novo-usuario2",
    "order": "created_at.asc"
})

logs_by_cadastro = defaultdict(list)
for l in logs:
    rb = l.get("request_body") or {}
    cid = rb.get("cadastro_id")
    if cid is None:
        continue
    cid = str(cid)
    if cid in cad_ids:
        logs_by_cadastro[cid].append(l)

def is_filled(v):
    if v is None:
        return False
    s = str(v).strip()
    return s != "" and s.lower() not in ("null", "none")

rows = []
not_found = []
status_enviado_sem_log = []

for cpf in cpfs:
    cads = cad_by_cpf.get(cpf, [])
    if not cads:
        not_found.append(cpf)
        rows.append({
            "cpf": cpf,
            "data_primeiro_cadastro": None,
            "status_primeiro": None,
            "cep_no_cadastro": None,
            "cep_no_payload": None,
            "qtd_chamadas_erp": 0,
            "qtd_com_cep": 0,
            "qtd_sem_cep": 0,
            "diagnostico": "NAO_ENVIADO_AINDA"
        })
        continue

    cads = sorted(cads, key=lambda x: x.get("created_at") or "")
    first = cads[0]
    endereco = first.get("endereco") if isinstance(first.get("endereco"), dict) else {}
    payload = first.get("payload_erp") if isinstance(first.get("payload_erp"), dict) else {}
    cep_cadastro = endereco.get("cep")
    cep_payload = payload.get("dados", {}).get("responsavelFinanceiro", {}).get("endereco", {}).get("cep")

    all_logs = []
    for c in cads:
        all_logs.extend(logs_by_cadastro.get(str(c.get("id")), []))
    all_logs = list({str(l.get("id")): l for l in all_logs}.values())

    com_cep = 0
    sem_cep = 0
    for l in all_logs:
        rb = l.get("request_body") if isinstance(l.get("request_body"), dict) else {}
        cep_req = rb.get("dados", {}).get("responsavelFinanceiro", {}).get("endereco", {}).get("cep")
        if is_filled(cep_req):
            com_cep += 1
        else:
            sem_cep += 1

    if len(all_logs) == 0:
        diag = "NAO_ENVIADO_AINDA"
    elif com_cep > 0:
        diag = "ENVIADO_COM_CEP"
    else:
        diag = "ENVIADO_SEM_CEP"

    if any((str(c.get("status") or "").strip().lower() == "enviado") for c in cads) and len(all_logs) == 0:
        status_enviado_sem_log.append(cpf)

    rows.append({
        "cpf": cpf,
        "data_primeiro_cadastro": first.get("created_at"),
        "status_primeiro": first.get("status"),
        "cep_no_cadastro": cep_cadastro,
        "cep_no_payload": cep_payload,
        "qtd_chamadas_erp": len(all_logs),
        "qtd_com_cep": com_cep,
        "qtd_sem_cep": sem_cep,
        "diagnostico": diag
    })

cpfs_total = len(cpfs)
cpfs_com_cep_no_cadastro = sum(1 for r in rows if is_filled(r["cep_no_cadastro"]))
cpfs_sem_cep_no_cadastro = cpfs_total - cpfs_com_cep_no_cadastro
cpfs_com_envio_com_cep = sum(1 for r in rows if r["diagnostico"] == "ENVIADO_COM_CEP")
cpfs_so_envio_sem_cep = sum(1 for r in rows if r["diagnostico"] == "ENVIADO_SEM_CEP")
cpfs_sem_chamada_erp = sum(1 for r in rows if r["diagnostico"] == "NAO_ENVIADO_AINDA")

out = {
    "cpf_count_raw": len(raw_cpfs),
    "cpf_count_dedup": len(cpfs),
    "detalhado": rows,
    "resumo": {
        "cpfs_total": cpfs_total,
        "cpfs_com_cep_no_cadastro": cpfs_com_cep_no_cadastro,
        "cpfs_sem_cep_no_cadastro": cpfs_sem_cep_no_cadastro,
        "cpfs_com_envio_com_cep": cpfs_com_envio_com_cep,
        "cpfs_so_envio_sem_cep": cpfs_so_envio_sem_cep,
        "cpfs_sem_chamada_erp": cpfs_sem_chamada_erp
    },
    "excecoes": {
        "cpfs_nao_encontrados_em_cadastros": not_found,
        "cpfs_status_enviado_sem_chamada_api_logs": sorted(set(status_enviado_sem_log))
    },
    "amostras": {
        "enviado_com_cep": [r for r in rows if r["diagnostico"] == "ENVIADO_COM_CEP"][:5],
        "enviado_sem_cep": [r for r in rows if r["diagnostico"] == "ENVIADO_SEM_CEP"][:5],
        "nao_enviado": [r for r in rows if r["diagnostico"] == "NAO_ENVIADO_AINDA"][:5]
    },
    "meta": {
        "cadastros_encontrados": len(cadastros),
        "api_logs_endpoint_erp_novo_usuario2_lidos": len(logs),
        "api_logs_vinculados_aos_cadastros_alvo": sum(len(v) for v in logs_by_cadastro.values())
    }
}

print(json.dumps(out, ensure_ascii=False))

